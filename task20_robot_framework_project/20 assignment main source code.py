# ============================================================
# TASK 20 - ROBOT FRAMEWORK STYLE AUTOMATION 
# Application : https://www.saucedemo.com/
# Framework    : Selenium Python
# Test By     : Ritik Kumar Singh
# ============================================================

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

# ============================================================
# CREATE EXCEL REPORT
# ============================================================

workbook = Workbook()
sheet = workbook.active
sheet.title = "Robot_Framework_Report"

headers = [
    "Test Case ID",
    "Scenario",
    "Expected Result",
    "Actual Result",
    "Status",
    "Execution Date"
]

for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)

# ============================================================
# BASE CLASS
# ============================================================

class SauceDemoAutomation:

    def __init__(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")

        self.driver = webdriver.Chrome(options=options)
        self.wait = WebDriverWait(self.driver, 10)

    # --------------------------------------------------------
    # OPEN WEBSITE
    # --------------------------------------------------------
    def open_website(self):
        self.driver.get("https://www.saucedemo.com/")

    # --------------------------------------------------------
    # LOGIN METHOD
    # --------------------------------------------------------
    def login(self, username, password):

        self.wait.until(
            EC.visibility_of_element_located((By.ID, "user-name"))
        ).send_keys(username)

        self.driver.find_element(By.ID, "password").send_keys(password)

        self.driver.find_element(By.ID, "login-button").click()

    # --------------------------------------------------------
    # VALID LOGIN TEST
    # --------------------------------------------------------
    def valid_login_test(self):

        self.open_website()

        self.login("standard_user", "secret_sauce")

        try:
            products_text = self.wait.until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//span[text()='Products']")
                )
            )

            self.write_result(
                "TC_001",
                "Valid Login",
                "Products page should open",
                "Products page opened",
                "PASS"
            )

            print("Valid Login Test Passed")

        except TimeoutException:

            self.write_result(
                "TC_001",
                "Valid Login",
                "Products page should open",
                "Products page not opened",
                "FAIL"
            )

            print("Valid Login Test Failed")

    # --------------------------------------------------------
    # INVALID LOGIN TEST
    # --------------------------------------------------------
    def invalid_login_test(self):

        self.open_website()

        self.login("invalid_user", "wrongpass")

        try:
            error_message = self.wait.until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//h3[contains(text(),'Epic sadface')]")
                )
            )

            self.write_result(
                "TC_002",
                "Invalid Login",
                "Error message should display",
                "Error message displayed",
                "PASS"
            )

            print("Invalid Login Test Passed")

        except TimeoutException:

            self.write_result(
                "TC_002",
                "Invalid Login",
                "Error message should display",
                "Error message not displayed",
                "FAIL"
            )

            print("Invalid Login Test Failed")

    # --------------------------------------------------------
    # ADD PRODUCT TO CART TEST
    # --------------------------------------------------------
    def add_product_test(self):

        self.open_website()

        self.login("standard_user", "secret_sauce")

        self.wait.until(
            EC.element_to_be_clickable(
                (By.ID, "add-to-cart-sauce-labs-backpack")
            )
        ).click()

        self.driver.find_element(By.CLASS_NAME, "shopping_cart_link").click()

        try:
            product = self.wait.until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//div[text()='Sauce Labs Backpack']")
                )
            )

            self.write_result(
                "TC_003",
                "Add Product To Cart",
                "Product should display in cart",
                "Product displayed in cart",
                "PASS"
            )

            print("Add Product Test Passed")

        except TimeoutException:

            self.write_result(
                "TC_003",
                "Add Product To Cart",
                "Product should display in cart",
                "Product not displayed",
                "FAIL"
            )

            print("Add Product Test Failed")

    # --------------------------------------------------------
    # CHECKOUT TEST
    # --------------------------------------------------------
    def checkout_test(self):

        self.open_website()

        self.login("standard_user", "secret_sauce")

        self.wait.until(
            EC.element_to_be_clickable(
                (By.ID, "add-to-cart-sauce-labs-backpack")
            )
        ).click()

        self.driver.find_element(
            By.ID,
            "add-to-cart-sauce-labs-bike-light"
        ).click()

        self.driver.find_element(
            By.CLASS_NAME,
            "shopping_cart_link"
        ).click()

        self.driver.find_element(By.ID, "checkout").click()

        self.driver.find_element(By.ID, "first-name").send_keys("Ritik")
        self.driver.find_element(By.ID, "last-name").send_keys("Tomer")
        self.driver.find_element(By.ID, "postal-code").send_keys("201001")

        self.driver.find_element(By.ID, "continue").click()

        try:
            checkout_page = self.wait.until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//span[text()='Checkout: Overview']")
                )
            )

            self.write_result(
                "TC_004",
                "Checkout Validation",
                "Checkout page should display",
                "Checkout page displayed",
                "PASS"
            )

            print("Checkout Test Passed")

        except TimeoutException:

            self.write_result(
                "TC_004",
                "Checkout Validation",
                "Checkout page should display",
                "Checkout page not displayed",
                "FAIL"
            )

            print("Checkout Test Failed")

    # --------------------------------------------------------
    # WRITE RESULTS TO EXCEL
    # --------------------------------------------------------
    def write_result(self, tc_id, scenario,
                     expected, actual, status):

        row = sheet.max_row + 1

        sheet.cell(row=row, column=1, value=tc_id)
        sheet.cell(row=row, column=2, value=scenario)
        sheet.cell(row=row, column=3, value=expected)
        sheet.cell(row=row, column=4, value=actual)
        sheet.cell(row=row, column=5, value=status)
        sheet.cell(
            row=row,
            column=6,
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )

    # --------------------------------------------------------
    # CLOSE BROWSER
    # --------------------------------------------------------
    def close_browser(self):
        self.driver.quit()


# ============================================================
# MAIN EXECUTION
# ============================================================

if __name__ == "__main__":

    test = SauceDemoAutomation()

    # Execute Test Cases
    test.valid_login_test()

    test.invalid_login_test()

    test.add_product_test()

    test.checkout_test()

    # Save Excel Report
    workbook.save("Robot_Framework_Test_Report.xlsx")

    # Close Browser
    test.close_browser()

    print("================================================")
    print("ALL TEST CASES EXECUTED SUCCESSFULLY")
    print("Excel Report Generated Successfully")
    print("================================================")

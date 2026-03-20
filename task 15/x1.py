# test_login_

import pytest
import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# CONFIG
# =======================
URL = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
EXCEL_FILE = "testdata.xlsx"

# =======================
# EXCEL FUNCTIONS
# =======================
def read_data():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    data = []

    for row in range(2, sheet.max_row + 1):
        username = sheet.cell(row, 2).value
        password = sheet.cell(row, 3).value
        data.append((row, username, password))

    return data

def write_result(row, result):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    sheet.cell(row, 7).value = result
    sheet.cell(row, 4).value = datetime.now().strftime("%Y-%m-%d")
    sheet.cell(row, 5).value = datetime.now().strftime("%H:%M:%S")

    wb.save(EXCEL_FILE)

# =======================
# PAGE OBJECT MODEL

class LoginPage:

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 10)

    username = (By.NAME, "username")
    password = (By.NAME, "password")
    login_btn = (By.XPATH, "//button[@type='submit']")
    dashboard = (By.XPATH, "//h6[text()='Dashboard']")

    def login(self, user, pwd):
        self.wait.until(EC.visibility_of_element_located(self.username)).clear()
        self.driver.find_element(*self.username).send_keys(user)

        self.driver.find_element(*self.password).clear()
        self.driver.find_element(*self.password).send_keys(pwd)

        self.driver.find_element(*self.login_btn).click()

    def is_login_successful(self):
        try:
            self.wait.until(EC.visibility_of_element_located(self.dashboard))
            return True
        except:
            return False

# =======================
# PYTEST FIXTURE
# =======================
@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(URL)
    yield driver
    driver.quit()

# =======================
# TEST CASE (DDTF + POM)
# =======================
@pytest.mark.parametrize("row,username,password", read_data())
def test_login(driver, row, username, password):

    login_page = LoginPage(driver)
    login_page.login(username, password)

    if login_page.is_login_successful():
        write_result(row, "Pass")
        assert True
    else:
        write_result(row, "Fail")
        assert False